import openpyxl


def dp1():
    list = []
    path = 'Appium_Python\\TestData\\testdata.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name('Sheet1')

    rowcount = sheet.max_row
    col_count = sheet.max_column

    # print(rowcount, col_count)

    for r in range(2, rowcount + 1):
        user = sheet.cell(r, 1).value
        pass1 = sheet.cell(r, 2).value

        tup = (user, pass1)
        list.append(tup)

    return list


method_name_start_row = 0
test_data_ends = 0
test_col_ends = 0


def dp2(method_name):
    list = []
    path = 'D:\\PythonFolder\\pythonProject\\pytestDemoForTesting\\testdata.xlsx'
    path = './TestData/testdata.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name('Sheet1')

    rowcount = sheet.max_row
    col_count = sheet.max_column
    print('max col', col_count)
    global method_from_sheet
    global method_name_start_row
    global test_data_ends
    global test_col_ends
    flag1 = True
    for r in range(1, rowcount + 1):
        # user = sheet.cell(r, 1).value
        # pass1 = sheet.cell(r, 2).value
        #
        # tup = (user, pass1)
        # list.append(tup)

        method_from_sheet = sheet.cell(r, 1).value
        if method_from_sheet == method_name:
            method_name_start_row = r
            print('Method starts from row: ', method_name_start_row)
            flag1 = False
            break
    if flag1:
        return
    method_name_start_row = method_name_start_row + 2

    # This will give the number of cols for a test row
    for c1 in range(1, col_count + 2):
        text2 = sheet.cell(method_name_start_row - 1, c1).value
        print(type(text2))
        print(text2)
        if text2 is None:
            print('Value of c1', c1)
            test_col_ends = c1
            break
    print('testcolends', test_col_ends)
    print('Max no of row is', rowcount)

    # this will give end of the test case row
    for r1 in range(method_name_start_row, rowcount + 2):
        text1 = sheet.cell(r1, 1).value
        print('text1 from testdata is ', text1)
        if text1 is None:
            print('row is ', text1)
            test_data_ends = r1 - 1
            break
    print('test data end row is', test_data_ends)
    # This will iterate  to get the test data
    for r2 in range(method_name_start_row, test_data_ends + 1):
        l1 = []
        for c3 in range(1, test_col_ends):
            # d1 = sheet.cell(r2, c3).value
            # pass1 = sheet.cell(r2, 2).value
            print(sheet.cell(r2, c3).value)
            l1.append(sheet.cell(r2, c3).value)
            print('test_col_ends', test_col_ends)

        tup = tuple(l1)
        # tup = (username, pass1)
        list.append(tup)
    list2 = []
    if len(list[0]) == 1:
        for i in range(0, len(list)):
            tup2 = list[i]
            val1 = tup2[0]
            list2.append(val1)
        return list2
    return list
